import json

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import requests

distance_matrix = []


class Node():
    def __init__(self):
        self.x = None
        self.y = None
        self.truck_num = 0


# 读取工作簿和工作簿中的工作表
def read_data(node_num):
    node = []
    data_frame = pd.read_excel('./昆明网点地址库V1-2022.10.14.xlsx', sheet_name='Sheet4')
    for i in range(node_num):
        n = Node()
        n.x = data_frame.iloc[2 + i, 3]
        n.y = data_frame.iloc[2 + i, 4]
        n.truck_num = data_frame.iloc[2 + i, 7] + data_frame.iloc[2 + i, 8] + data_frame.iloc[2 + i, 9]
        node.append(n)
    global distance_matrix
    for i in node:
        print('正在收集'+str(node.index(i))+'到其他点的距离')
        temp = []
        for j in node:
            temp.append(get_distance(i, j))
        distance_matrix.append(temp)
    return node


def gethtml(url):
    i = 0
    while i < 3:
        try:
            html = requests.get(url, timeout=5).text
            return html
        except requests.exceptions.RequestException:
            i += 1


def get_distance(i, j):
    #print(i.x, i.y, j.x, j.y)
    url = 'https://api.map.baidu.com/direction/v2/driving?origin=' + str(i.y) + ',' + str(
        i.x) + '&destination=' + str(j.y) + ',' + str(j.x) + '&tactics=2&ak=yourAK'
    a=gethtml(url)
    data = json.loads(a)
    #print(data)
    return data['result']['routes'][0]['distance']
    # return np.hypot((i.x-j.x),(i.y-j.y))


def cal_sol(node):
    distance_row = []
    for i, ii in enumerate(node):
        temp = 0
        for j, jj in enumerate(node):
            temp += distance_matrix[i][j] * jj.truck_num
        distance_row.append(temp)
    return node[distance_row.index(min(distance_row))]


def write_excel(best_node, node):
    book = openpyxl.load_workbook(r'./昆明网点地址库V1-2022.10.14.xlsx')
    # 根据名称获取某个sheet对象
    # print(book.sheetnames)
    sh = book['Sheet4']

    for i in range(len(node)):
        if node.index(best_node)!=i:
            sh.cell(4 + i, 6).value = str(distance_matrix[node.index(best_node)][i])
        else:
            sh.cell(4 + i, 6).value = 0
        # data_frame.loc[2+i][5]=get_distance(best_node,node[i])
    sh.cell(4, 17).value = best_node.x
    sh.cell(5, 17).value = best_node.y
    book.save('昆明网点地址库V1-2022.10.14.xlsx')


# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    node_num = 12  # 代入计算节点个数
    print('=======获取数据中=======')
    node = read_data(node_num)
    print('=======获取数据完毕=======')
    best_node = cal_sol(node)
    write_excel(best_node, node)
    '''
    for i in node:
        if i != best_node:
            plt.scatter(i.x, i.y, c='g')
            plt.plot([best_node.x,i.x],[best_node.y,i.y],c='b')
        else:
            plt.scatter(i.x, i.y, c='r')
    plt.show()
    '''
